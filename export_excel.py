"""
导出Excel报告工具

功能：将参数对比结果导出为Excel表格，方便工程师阅读

使用方法：
1. 先运行 extract_params.py 提取参数
2. 再运行 compare_params.py 进行对比
3. 设置 COMPARISON_RESULT 为对比结果文件路径
4. 运行脚本：python export_excel.py
"""

import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# 配置区域
# ============================================================

# 对比结果文件路径（修改为实际文件名）
COMPARISON_RESULT = "output/com_azure_01_28_1144.json"

# 提取结果文件路径（用于获取未匹配参数的完整信息）
EXTRACTION_RESULT = "output/ex_azure_01_28_1057.json"

# 输出目录
OUTPUT_DIR = "output"


# ============================================================
# 样式定义
# ============================================================

# 表头样式
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 数据行样式
DATA_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# 合规/不合规样式
COMPLIANT_FONT = Font(color="008000", bold=True)  # 绿色
NON_COMPLIANT_FONT = Font(color="FF0000", bold=True)  # 红色
NO_MATCH_FONT = Font(color="808080")  # 灰色

# 分隔行样式
SEPARATOR_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
SEPARATOR_FONT = Font(bold=True, color="806000")

# 边框
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 类型颜色
TYPE_FILLS = {
    "A": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),  # 浅红
    "B": PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid"),  # 浅橙
    "C": PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"),  # 浅绿
    "D": PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),  # 浅蓝
}


def load_json(file_path: str) -> dict:
    """加载JSON文件"""
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)


def export_to_excel(comparison_result: dict, extraction_result: dict, output_path: str):
    """导出对比结果到Excel"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "参数对比结果"
    
    # 设置列宽
    column_widths = [35, 18, 35, 18, 12, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 写入表头
    headers = ["用户参数", "用户值", "匹配规范参数", "规范值", "参数类型", "是否合规"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    current_row = 2
    
    # 1. 写入符合规范的参数
    compliant_params = comparison_result.get("compliant_params", [])
    if compliant_params:
        for item in compliant_params:
            current_row = write_param_row(ws, current_row, item, "compliant")
    
    # 2. 写入不符合规范的参数
    non_compliant_params = comparison_result.get("non_compliant_params", [])
    if non_compliant_params:
        for item in non_compliant_params:
            current_row = write_param_row(ws, current_row, item, "non_compliant")
    
    # 3. 写入无法判断的参数
    uncertain_params = comparison_result.get("uncertain_params", [])
    if uncertain_params:
        for item in uncertain_params:
            current_row = write_param_row(ws, current_row, item, "uncertain")
    
    # 4. 写入未匹配到规范的参数
    no_match_params = comparison_result.get("no_match_params", [])
    if no_match_params:
        for item in no_match_params:
            current_row = write_param_row(ws, current_row, item, "no_match")
    
    # 5. 添加分隔行 - 未提取到的参数
    not_found_params = extraction_result.get("not_found", [])
    if not_found_params:
        # 添加空行
        current_row += 1
        
        # 添加分隔标题行
        separator_cell = ws.cell(row=current_row, column=1, value="以下参数在文档中未提取到")
        separator_cell.font = SEPARATOR_FONT
        separator_cell.fill = SEPARATOR_FILL
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        for col in range(1, 7):
            ws.cell(row=current_row, column=col).border = THIN_BORDER
            ws.cell(row=current_row, column=col).fill = SEPARATOR_FILL
        current_row += 1
        
        # 写入未提取到的参数
        for param_name in not_found_params:
            ws.cell(row=current_row, column=1, value=param_name).alignment = DATA_ALIGNMENT
            ws.cell(row=current_row, column=2, value="-").alignment = CENTER_ALIGNMENT
            ws.cell(row=current_row, column=3, value="-").alignment = CENTER_ALIGNMENT
            ws.cell(row=current_row, column=4, value="-").alignment = CENTER_ALIGNMENT
            ws.cell(row=current_row, column=5, value="-").alignment = CENTER_ALIGNMENT
            ws.cell(row=current_row, column=6, value="-").alignment = CENTER_ALIGNMENT
            
            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.border = THIN_BORDER
                cell.font = NO_MATCH_FONT
            
            current_row += 1
    
    # 添加统计信息sheet
    add_statistics_sheet(wb, comparison_result, extraction_result)
    
    # 保存文件
    wb.save(output_path)
    print(f"✅ Excel报告已保存到: {output_path}")


def write_param_row(ws, row: int, item: dict, status: str) -> int:
    """写入一行参数数据"""
    
    user_name = item.get("user_param_name", "")
    user_value = item.get("user_value", "")
    spec_name = item.get("matched_spec_name", "")
    spec_value = item.get("spec_value", "")
    param_type = item.get("param_type", "")
    is_compliant = item.get("is_compliant")
    
    # 写入数据
    ws.cell(row=row, column=1, value=user_name).alignment = DATA_ALIGNMENT
    ws.cell(row=row, column=2, value=user_value).alignment = CENTER_ALIGNMENT
    ws.cell(row=row, column=3, value=spec_name if spec_name else "-").alignment = DATA_ALIGNMENT
    ws.cell(row=row, column=4, value=spec_value if spec_value else "-").alignment = CENTER_ALIGNMENT
    ws.cell(row=row, column=5, value=param_type if param_type else "-").alignment = CENTER_ALIGNMENT
    
    # 是否合规
    if status == "compliant":
        compliance_cell = ws.cell(row=row, column=6, value="✓")
        compliance_cell.font = COMPLIANT_FONT
    elif status == "non_compliant":
        compliance_cell = ws.cell(row=row, column=6, value="✗")
        compliance_cell.font = NON_COMPLIANT_FONT
    elif status == "no_match":
        compliance_cell = ws.cell(row=row, column=6, value="-")
        compliance_cell.font = NO_MATCH_FONT
    else:
        compliance_cell = ws.cell(row=row, column=6, value="?")
    
    compliance_cell.alignment = CENTER_ALIGNMENT
    
    # 应用边框和类型颜色
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER
    
    # 类型列着色
    if param_type in TYPE_FILLS:
        ws.cell(row=row, column=5).fill = TYPE_FILLS[param_type]
    
    return row + 1


def add_statistics_sheet(wb: Workbook, comparison_result: dict, extraction_result: dict):
    """添加统计信息sheet"""
    
    ws = wb.create_sheet(title="统计信息")
    
    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    
    stats = comparison_result.get("statistics", {})
    type_stats = comparison_result.get("type_statistics", {})
    type_defs = comparison_result.get("type_definitions", {})
    extraction_stats = extraction_result.get("statistics", {})
    
    row = 1
    
    # 标题
    ws.cell(row=row, column=1, value="参数对比统计报告").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 2
    
    # 提取统计
    ws.cell(row=row, column=1, value="提取统计").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="预定义参数总数")
    ws.cell(row=row, column=2, value=extraction_stats.get("total_requested", 0))
    row += 1
    ws.cell(row=row, column=1, value="成功提取")
    ws.cell(row=row, column=2, value=extraction_stats.get("found", 0))
    row += 1
    ws.cell(row=row, column=1, value="未提取到")
    ws.cell(row=row, column=2, value=extraction_stats.get("not_found", 0))
    row += 2
    
    # 对比统计
    ws.cell(row=row, column=1, value="对比统计").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="参与对比总数")
    ws.cell(row=row, column=2, value=stats.get("total", 0))
    row += 1
    ws.cell(row=row, column=1, value="符合规范")
    ws.cell(row=row, column=2, value=stats.get("compliant", 0))
    ws.cell(row=row, column=2).font = COMPLIANT_FONT
    row += 1
    ws.cell(row=row, column=1, value="不符合规范")
    ws.cell(row=row, column=2, value=stats.get("non_compliant", 0))
    ws.cell(row=row, column=2).font = NON_COMPLIANT_FONT
    row += 1
    ws.cell(row=row, column=1, value="未匹配到规范")
    ws.cell(row=row, column=2, value=stats.get("no_match", 0))
    row += 1
    ws.cell(row=row, column=1, value="无法判断")
    ws.cell(row=row, column=2, value=stats.get("uncertain", 0))
    row += 2
    
    # 按类型统计
    ws.cell(row=row, column=1, value="按类型统计").font = Font(bold=True, size=12)
    row += 1
    
    for ptype in ["A", "B", "C", "D"]:
        type_data = type_stats.get(ptype, {})
        type_def = type_defs.get(ptype, "")
        ws.cell(row=row, column=1, value=f"{ptype}类 ({type_def})")
        compliant = type_data.get("compliant", 0)
        non_compliant = type_data.get("non_compliant", 0)
        ws.cell(row=row, column=2, value=f"符合:{compliant} 不符合:{non_compliant}")
        if ptype in TYPE_FILLS:
            ws.cell(row=row, column=1).fill = TYPE_FILLS[ptype]
        row += 1
    
    row += 1
    
    # 文件信息
    ws.cell(row=row, column=1, value="文件信息").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="提取结果文件")
    ws.cell(row=row, column=2, value=comparison_result.get("extraction_file", ""))
    row += 1
    ws.cell(row=row, column=1, value="规范数据库")
    ws.cell(row=row, column=2, value=comparison_result.get("spec_database", ""))
    row += 1
    ws.cell(row=row, column=1, value="对比时间")
    ws.cell(row=row, column=2, value=comparison_result.get("compare_time", ""))
    row += 1
    ws.cell(row=row, column=1, value="使用模型")
    ws.cell(row=row, column=2, value=comparison_result.get("model", ""))


def main():
    """主函数"""
    print(f"\n{'='*60}")
    print(f"导出Excel报告工具")
    print(f"{'='*60}")
    print(f"对比结果: {COMPARISON_RESULT}")
    print(f"提取结果: {EXTRACTION_RESULT}")
    
    # 检查文件
    if not os.path.exists(COMPARISON_RESULT):
        print(f"\n❌ 对比结果文件不存在: {COMPARISON_RESULT}")
        print(f"请先运行 compare_params.py 进行对比")
        return
    
    if not os.path.exists(EXTRACTION_RESULT):
        print(f"\n❌ 提取结果文件不存在: {EXTRACTION_RESULT}")
        print(f"请先运行 extract_params.py 进行提取")
        return
    
    # 加载数据
    print(f"\n📂 加载数据...")
    comparison_result = load_json(COMPARISON_RESULT)
    extraction_result = load_json(EXTRACTION_RESULT)
    
    print(f"  ✓ 对比结果: {comparison_result.get('statistics', {}).get('total', 0)} 个参数")
    print(f"  ✓ 未提取到: {len(extraction_result.get('not_found', []))} 个参数")
    
    # 生成输出文件名
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%m_%d_%H%M")
    output_file = os.path.join(OUTPUT_DIR, f"report_azure_{timestamp}.xlsx")
    
    # 导出Excel
    print(f"\n📊 生成Excel报告...")
    export_to_excel(comparison_result, extraction_result, output_file)
    
    # 统计信息
    stats = comparison_result.get("statistics", {})
    print(f"\n📈 统计摘要:")
    print(f"  符合规范:   {stats.get('compliant', 0)}")
    print(f"  不符合规范: {stats.get('non_compliant', 0)}")
    print(f"  未匹配规范: {stats.get('no_match', 0)}")
    print(f"  未提取到:   {len(extraction_result.get('not_found', []))}")


if __name__ == "__main__":
    main()
